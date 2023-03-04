import json
import openpyxl
import datetime
import collections
import firebase_admin
from firebase_admin import firestore
from firebase_admin import credentials
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side


def fetch_ref(collection):
    cred = credentials.Certificate('py/credentials.json')
    firebase_admin.initialize_app(cred)

    db = firestore.client()
    return db.collection(collection)


def main():
    date = datetime.datetime.now()
    headers = ['{0:%Y/%m/%d %H:%M更新} (自動更新)'.format(date), '時間', '名前', '学年', 'メールアドレス', '備考', '受付完了時刻']
    titles = ['スライム', 'DNAストラップ', 'くるくるマグネット', 'ふしぎな水そう', '球体ロボットでプログラミング体験', '電卓を分解して自分だけのアクセサリーを作ろう']
    keyOrder = ['title', 'schedule', 'name', 'grade', 'email', 'memo', 'submitted_at']

    # collectionの取得
    applicants_ref = fetch_ref('door-applicants')

    # フォント・罫線・カラー
    font = Font(name='游ゴシック')
    thin = Side(style='thin', color='000000')
    dashed = Side(style='dashed', color='000000')
    fill = PatternFill(patternType='solid', fgColor='ebf1de')

    # Excelファイルの初期化
    wb = openpyxl.Workbook()

    # データの書き込み
    for index, title in enumerate(titles):
        # シートの追加
        wb.create_sheet(index=index, title=title)
        ws = wb[title]

        # ヘッダーの書き込み
        for col, header in enumerate(headers):
            cell = ws.cell(row=1, column=col+1,value=header)
            cell.font = font
            cell.fill = fill
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        # DBにアクセス
        docs = applicants_ref.document(title).get().to_dict()
        ordered_docs = collections.OrderedDict(sorted(docs.items()))
        
        row = 2
        title_written = False        
        for schedule, users in ordered_docs.items():
            if not users:
                continue

            schedule_written = False

            for user in users:
                if title_written:
                    title = ''
                else:
                    title_written = True

                if schedule_written:
                    schedule = ''
                else:
                    schedule_written = True
                
                user['title'], user['schedule'] = title, schedule
                for col, key in enumerate(keyOrder):
                    border = Border(top=thin, bottom=thin, left=dashed, right=dashed)

                    if key == 'title' or key == 'schedule':
                        border = Border(left=thin, right=thin)
                    elif key == 'name':
                        border = Border(top=thin, bottom=thin, left=thin, right=dashed)
                    elif key == 'submitted_at':
                        user[key] = str(user[key])
                        border = Border(top=thin, bottom=thin, left=dashed, right=thin)

                    cell = ws.cell(row=row, column=col+1, value=user[key])
                    cell.font = font
                    cell.border = border
                
                row += 1
            
            ws.cell(row=row-1, column=2).border = Border(bottom=thin, left=thin, right=dashed)

        ws.cell(row=row-1, column=1).border = Border(bottom=thin, left=thin, right=thin)

    # 保存
    wb.save('py/applicants/当日申込者リスト.xlsx')


if __name__ == '__main__':
    main()